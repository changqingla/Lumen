#
#  Copyright 2025 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

"""
智能表格解析器（带合并单元格自动处理）

这个解析器是 table.py 的简化版本，主要特性：
1. 自动检测和处理合并单元格
2. 支持标题行保护
3. 适配不规范的表格格式
4. 简化的分块策略：只生成全文内容，不做复杂的数据类型识别

与 table.py 的区别：
- table.py: 
  * 要求表格规范，不处理合并单元格
  * 复杂的数据类型识别和结构化字段生成
  * 生成拼音字段名 (如: ming_ye_tks, nian_ling_long)
  
- ir-table.py: 
  * 自动处理合并单元格，容错性更强
  * 简化的分块逻辑：只生成全文内容 (content_ltks)
  * 每行数据格式: "列名:值; 列名:值; ..."
  * 不做数据类型识别，直接使用原始值
"""

import logging
import re
from io import BytesIO
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from deepdoc.parser.utils import get_text
from core.nlp import rag_tokenizer, tokenize
from deepdoc.parser import ExcelParser

logger = logging.getLogger(__name__)


def unmerge_and_fill_excel(file_path, keep_title=True, unmerge_start_row=2, 
                           unmerge_end_row=None, only_columns=None, 
                           exclude_sheets=None, include_sheets=None):
    """
    智能处理 Excel 所有工作表的合并单元格，使其符合 table 解析器要求
    
    Args:
        file_path (str): Excel 文件路径（可以是文件路径或二进制数据）
        keep_title (bool): 是否保留第1行的合并单元格（默认 True）
        unmerge_start_row (int): 从哪一行开始取消合并（默认 2，即从第2行开始）
        unmerge_end_row (int): 到哪一行结束（默认 None，处理到最后一行）
        only_columns (list): 只处理指定列（如 ["A", "B"]，默认 None 处理所有列）
        exclude_sheets (list): 排除某些工作表（默认 None）
        include_sheets (list): 只处理某些工作表（默认 None，处理所有）
        
    Returns:
        BytesIO: 处理后的 Excel 文件二进制数据
        
    Example:
        >>> processed = unmerge_and_fill_excel(
        ...     "城市管理标准清单.xlsx",
        ...     keep_title=True,
        ...     unmerge_start_row=3
        ... )
    """
    # 加载工作簿
    if isinstance(file_path, (str, Path)):
        wb = load_workbook(file_path)
    else:
        # 如果是二进制数据
        wb = load_workbook(BytesIO(file_path))
    
    # 确定要处理的工作表
    all_sheets = wb.sheetnames
    
    if include_sheets is not None:
        # 白名单模式：只处理指定的工作表
        sheets_to_process = [s for s in all_sheets if s in include_sheets]
    elif exclude_sheets is not None:
        # 黑名单模式：排除指定的工作表
        sheets_to_process = [s for s in all_sheets if s not in exclude_sheets]
    else:
        # 默认：处理所有工作表
        sheets_to_process = all_sheets
    
    # 列名转换辅助函数
    def col_letter_to_index(col_letter):
        """将列字母转换为索引（A->1, B->2, ...）"""
        index = 0
        for char in col_letter:
            index = index * 26 + (ord(char.upper()) - ord('A') + 1)
        return index
    
    # 处理每个工作表
    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        
        # 获取所有合并单元格区域
        merged_ranges = list(ws.merged_cells.ranges)
        
        # 存储需要处理的合并单元格信息
        to_process = []
        
        for merged_range in merged_ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            
            # 判断1: 是否在第1行（标题行）
            if keep_title and min_row == 1 and max_row == 1:
                continue  # 跳过标题行的合并
            
            
            # 判断2: 是否在指定的取消合并行范围内
            if min_row < unmerge_start_row:
                continue  # 不在范围内
            
            if unmerge_end_row and max_row > unmerge_end_row:
                continue  # 超出范围
            
            # 判断4: 是否在指定的列范围内
            if only_columns is not None:
                # 检查合并区域的列是否在指定范围内
                col_in_range = False
                for col_letter in only_columns:
                    col_idx = col_letter_to_index(col_letter)
                    if min_col <= col_idx <= max_col:
                        col_in_range = True
                        break
                if not col_in_range:
                    continue
            
            # 通过所有判断，记录需要处理的合并单元格
            # 获取左上角单元格的值
            top_left_cell = ws.cell(row=min_row, column=min_col)
            value = top_left_cell.value
            
            to_process.append({
                'range': merged_range,
                'min_row': min_row,
                'max_row': max_row,
                'min_col': min_col,
                'max_col': max_col,
                'value': value
            })
        
        # 取消合并并填充值
        for item in to_process:
            # 取消合并
            ws.unmerge_cells(str(item['range']))
            
            # 填充值到所有单元格
            for row in range(item['min_row'], item['max_row'] + 1):
                for col in range(item['min_col'], item['max_col'] + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = item['value']
    
    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


class Excel(ExcelParser):
    def __call__(self, fnm, binary=None, from_page=0,
                 to_page=10000000000, callback=None, skip_first_row=False):
        """
        解析 Excel 文件
        
        Args:
            fnm: 文件名或路径
            binary: 二进制数据
            from_page: 起始行（数据行，不包括标题行）
            to_page: 结束行
            callback: 回调函数
            skip_first_row: 是否跳过第1行（默认 False）
                - True: 第1行为文档标题，第2行为列标题，第3+行为数据
                - False: 第1行为列标题，第2+行为数据
        
        Returns:
            tuple: (dataframes_list, sheet_info_dict)
                - dataframes_list: DataFrame 列表
                - sheet_info_dict: 工作表信息字典，包含每个工作表的标题等信息
        """
        if not binary:
            wb = Excel._load_excel_to_workbook(fnm)
        else:
            wb = Excel._load_excel_to_workbook(BytesIO(binary))
        total = 0
        for sheetname in wb.sheetnames:
            total += len(list(wb[sheetname].rows))

        res, fails, done = [], [], 0
        sheet_info = {}  # 存储每个工作表的信息
        rn = 0
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            
            # 根据 skip_first_row 参数决定列标题位置
            header_row_index = 1 if skip_first_row and len(rows) > 1 else 0
            data_start_index = header_row_index + 1
            
            # 如果跳过第1行，但总行数不足，跳过此工作表
            if skip_first_row and len(rows) <= 1:
                continue
            
            # 提取工作表标题（如果 skip_first_row=True，则第1行是标题）
            sheet_title = None
            if skip_first_row and len(rows) > 0:
                first_row = rows[0]
                title_parts = []
                for cell in first_row:
                    if cell.value and str(cell.value).strip():
                        title_parts.append(str(cell.value).strip())
                if title_parts:
                    sheet_title = " ".join(title_parts)
            
            # 存储工作表信息
            sheet_info[sheetname] = {
                'title': sheet_title,
                'sheet_name': sheetname
            }
            
            # 获取列标题（第2行或第1行）
            headers = [cell.value for cell in rows[header_row_index]]
            missed = set([i for i, h in enumerate(headers) if h is None])
            headers = [
                cell.value for i,
                cell in enumerate(
                    rows[header_row_index]) if i not in missed]
            if not headers:
                continue
            
            data = []
            # 从数据行开始读取（第3行或第2行）
            for i, r in enumerate(rows[data_start_index:]):
                rn += 1
                if rn - 1 < from_page:
                    continue
                if rn - 1 >= to_page:
                    break
                row = [
                    cell.value for ii,
                    cell in enumerate(r) if ii not in missed]
                if len(row) != len(headers):
                    fails.append(str(i))
                    continue
                data.append(row)
                done += 1
            if np.array(data).size == 0:
                continue
            res.append(pd.DataFrame(np.array(data), columns=headers))

        callback(0.3, ("Extract records: {}~{}".format(from_page + 1, min(to_page, from_page + rn)) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))
        return res, sheet_info


def chunk(filename, binary=None, from_page=0, to_page=10000000000,
          lang="Chinese", callback=None, **kwargs):
    """
    智能表格解析器（自动处理合并单元格）
    
    支持 Excel 和 CSV/TXT 格式文件。
    
    **自动处理功能**：
    - 自动检测并处理 Excel 中的合并单元格
    - 保留标题行格式（第1行）
    - 自动填充合并区域的空白单元格
    
    **表格要求**：
    - CSV/TXT: 列之间使用 TAB 分隔
    - 第一行或第二行必须是列标题
    - 列标题应该有意义，建议使用同义词/枚举值格式
    
    **列标题示例**：
        1. supplier/vendor\tcolor(yellow, red, brown)\tgender/sex(male, female)\tsize(M,L,XL,XXL)
        2. 姓名/名字\t电话/手机/微信\t最高学历（高中，职高，硕士，本科，博士，初中，中技，中专，专科，专升本，MPA，MBA，EMBA）
    
    **合并单元格处理参数**（通过 kwargs 传入）：
        - keep_title (bool): 是否保留第1行标题的合并（默认 True）
        - unmerge_start_row (int): 从哪一行开始取消合并（默认 2，包括列标题行）
        - unmerge_end_row (int): 到哪一行结束（默认 None，处理到最后）
        - only_columns (list): 只处理指定列（默认 None，处理所有列）
        - auto_unmerge (bool): 是否自动处理合并单元格（默认 True）
        
    **重要说明**：
        - 第1行（文档标题）：如果 keep_title=True，保持合并状态
        - 第2行（列标题）：会取消合并，允许重复的列名以保持表格结构一致
        - 第3+行（数据行）：取消合并并填充相同内容
    
    **输出格式**：
        每行数据会被转换为一个独立的 chunk
        
        chunk 结构：
        {
            "docnm_kwd": "文件名.xlsx",
            "title_tks": ["文件", "名"],
            "content_ltks": "标题:xxx; 列名1:值1; 列名2:值2; 列名3:值3",
            "content_with_weight": "标题:xxx; 列名1:值1; 列名2:值2; 列名3:值3",
            ... (其他 tokenize 生成的字段)
        }
        
        注意：
        - 不生成结构化字段（如 ming_ye_tks），只有全文内容
        - 如果 keep_title=True，每个 chunk 自动添加 "标题:xxx"
        - 标题来自 Excel 第1行的内容（合并所有非空单元格）
    
    Example:
        >>> chunks = chunk(
        ...     "城市管理标准清单.xlsx",
        ...     lang="Chinese",
        ...     keep_title=True,
        ...     unmerge_start_row=2  # 从第2行开始，包括列标题
        ... )
    """

    if re.search(r"\.xlsx?$", filename, re.IGNORECASE):
        callback(0.05, "Start to parse Excel file...")
        
        # 检查是否需要自动处理合并单元格（默认开启）
        auto_unmerge = kwargs.get("auto_unmerge", True)
        
        if auto_unmerge:
            callback(0.1, "Detecting and processing merged cells...")
            
            # 获取合并单元格处理参数
            keep_title = kwargs.get("keep_title", True)
            unmerge_start_row = kwargs.get("unmerge_start_row", 2)
            unmerge_end_row = kwargs.get("unmerge_end_row", None)
            only_columns = kwargs.get("only_columns", None)
            exclude_sheets = kwargs.get("exclude_sheets", None)
            include_sheets = kwargs.get("include_sheets", None)
            
            # 处理合并单元格
            try:
                processed_binary = unmerge_and_fill_excel(
                    file_path=binary if binary else filename,
                    keep_title=keep_title,
                    unmerge_start_row=unmerge_start_row,
                    unmerge_end_row=unmerge_end_row,
                    only_columns=only_columns,
                    exclude_sheets=exclude_sheets,
                    include_sheets=include_sheets
                )
                # 使用处理后的二进制数据
                binary = processed_binary.read()
                callback(0.2, "Merged cells processed successfully.")
            except Exception as e:
                # 如果处理失败，使用原始文件继续
                callback(0.2, f"Warning: Failed to process merged cells: {e}. Using original file.")
        
        callback(0.25, "Parsing Excel data...")
        excel_parser = Excel()
        # 只有当 keep_title=True 时才跳过第1行，使用第2行作为列名
        skip_first_row = kwargs.get("keep_title", True)
        dfs, sheet_info = excel_parser(
            filename,
            binary,
            from_page=from_page,
            to_page=to_page,
            callback=callback,
            skip_first_row=skip_first_row)
    elif re.search(r"\.(txt|csv)$", filename, re.IGNORECASE):
        callback(0.1, "Start to parse.")
        txt = get_text(filename, binary)
        lines = txt.split("\n")
        fails = []
        headers = lines[0].split(kwargs.get("delimiter", "\t"))
        rows = []
        for i, line in enumerate(lines[1:]):
            if i < from_page:
                continue
            if i >= to_page:
                break
            row = [field for field in line.split(kwargs.get("delimiter", "\t"))]
            if len(row) != len(headers):
                fails.append(str(i))
                continue
            rows.append(row)

        callback(0.3, ("Extract records: {}~{}".format(from_page, min(len(lines), to_page)) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))

        dfs = [pd.DataFrame(np.array(rows), columns=headers)]
        sheet_info = {}  # CSV/TXT 文件没有工作表信息

    else:
        raise NotImplementedError(
            "file type not supported yet(excel, text, csv supported)")

    res = []
    eng = lang.lower() == "english"
    
    # 为每个 DataFrame 匹配对应的工作表标题
    for df_index, df in enumerate(dfs):
        # 获取列名
        clmns = df.columns.values
        
        # 获取当前工作表的标题
        current_sheet_title = None
        if sheet_info:
            # 从 sheet_info 中获取对应工作表的标题
            sheet_names = list(sheet_info.keys())
            if df_index < len(sheet_names):
                sheet_name = sheet_names[df_index]
                current_sheet_title = sheet_info[sheet_name].get('title')
        
        # 逐行生成 chunk
        for ii, row in df.iterrows():
            # 基础元数据
            d = {
                "docnm_kwd": filename,
                "title_tks": rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", filename))
            }
            
            # 构建行文本：列名:值; 列名:值; ...
            row_txt = []
            
            # 如果有标题，先添加当前工作表的标题
            if current_sheet_title:
                row_txt.append(f"标题:{current_sheet_title}")
            
            # 处理列数据，正确处理重复列名
            # 获取唯一的列名，避免重复处理
            unique_col_names = []
            for col_name in clmns:
                if col_name not in unique_col_names:
                    unique_col_names.append(col_name)
            
            for col_name in unique_col_names:
                try:
                    value = row[col_name]
                    
                    # 跳过空值 - 安全的方式处理各种数据类型
                    if value is None:
                        continue
                    
                    # 检查是否是 pandas Series（重复列名的情况）
                    if hasattr(value, 'iloc') and hasattr(value, 'values'):
                        # 这是一个 Series，包含多个同名列的值
                        # 为每个非空值创建单独的 "列名:值" 条目
                        for val in value.values:
                            if val is not None and not pd.isna(val):
                                str_val = str(val).strip()
                                if str_val and str_val.lower() not in ['nan', 'none', 'null']:
                                    row_txt.append(f"{col_name}:{str_val}")
                        continue
                    
                    # 处理单个值的情况
                    try:
                        if pd.isna(value):
                            continue
                    except (TypeError, ValueError):
                        pass
                    
                    # 转换为字符串并检查是否为空
                    str_value = str(value).strip()
                    if not str_value or str_value.lower() in ['nan', 'none', 'null']:
                        continue
                    
                    # 添加到行文本
                    row_txt.append(f"{col_name}:{str_value}")
                    
                except Exception as exc:
                    # 如果处理某个字段时出错，跳过该字段但继续处理其他字段
                    logger.warning(
                        "Skipping an invalid table cell (error_type=%s)",
                        type(exc).__name__,
                    )
                    continue
            
            # 如果这行没有任何内容，跳过
            if not row_txt:
                continue
            
            # 生成全文内容（用分号分隔）
            full_text = "; ".join(row_txt)
            
            # 使用 tokenize 函数生成 content_ltks 等字段
            tokenize(d, full_text, eng)
            
            res.append(d)

    return res
