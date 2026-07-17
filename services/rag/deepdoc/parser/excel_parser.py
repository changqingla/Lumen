#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

"""
Excel文件解析器模块

该模块提供了DeepRAGExcelParser类，用于解析Excel文件（.xlsx, .xls）和CSV文件，
并将其转换为HTML表格或结构化文本格式，以便在RAG（检索增强生成）系统中使用。

主要功能：
1. 支持多种Excel格式的解析
2. 自动检测文件类型并选择合适的解析方法
3. 将Excel数据转换为HTML表格格式
4. 将Excel数据转换为结构化文本格式
5. 统计文件行数
"""

import logging
import sys
import os
from io import BytesIO

# 添加项目根目录到Python路径，以便正确导入rag模块
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.join(current_dir, '..', '..')
sys.path.insert(0, project_root)

import pandas as pd  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from core.nlp import find_codec  # noqa: E402


class DeepRAGExcelParser:
    """
    DeepRAG Excel文件解析器

    该类提供了解析Excel文件和CSV文件的功能，支持将文件内容转换为
    HTML表格格式或结构化文本格式，适用于文档检索和处理场景。
    """

    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        """
        将文件对象加载为openpyxl工作簿对象

        该方法能够智能识别文件类型，支持Excel文件和CSV文件的加载。
        对于Excel文件，优先使用openpyxl加载；对于CSV文件或加载失败的情况，
        使用pandas读取后转换为工作簿对象。

        Args:
            file_like_object: 文件对象，可以是字节数据、BytesIO对象或文件路径

        Returns:
            openpyxl.Workbook: 工作簿对象

        Raises:
            Exception: 当所有解析方法都失败时抛出异常
        """
        # 如果输入是字节数据，转换为BytesIO对象
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)

        # 如果输入是文件路径（字符串），直接尝试加载
        if isinstance(file_like_object, str):
            try:
                return load_workbook(file_like_object)
            except Exception as openpyxl_error:
                logging.info(
                    "Spreadsheet openpyxl path load failed; trying pandas: "
                    "error_type=%s",
                    type(openpyxl_error).__name__,
                )
                try:
                    df = pd.read_excel(file_like_object)
                    return DeepRAGExcelParser._dataframe_to_workbook(df)
                except Exception:
                    raise ValueError("Spreadsheet input could not be parsed") from None

        # 读取文件头部4个字节来判断文件类型
        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        # 检查是否为Excel文件格式
        # PK\x03\x04: .xlsx文件头标识
        # \xD0\xCF\x11\xE0: .xls文件头标识
        if not (file_head.startswith(b'PK\x03\x04') or file_head.startswith(b'\xD0\xCF\x11\xE0')):
            logging.info("****wxy: Not an Excel file, converting CSV to Excel Workbook")

            try:
                # 尝试作为CSV文件解析
                file_like_object.seek(0)
                df = pd.read_csv(file_like_object)
                return DeepRAGExcelParser._dataframe_to_workbook(df)

            except Exception:
                raise ValueError("Tabular input could not be parsed") from None

        # 尝试使用openpyxl加载Excel文件
        try:
            return load_workbook(file_like_object)
        except Exception as openpyxl_error:
            logging.info(
                "Spreadsheet openpyxl stream load failed; trying pandas: "
                "error_type=%s",
                type(openpyxl_error).__name__,
            )
            try:
                # openpyxl失败时，尝试使用pandas读取
                file_like_object.seek(0)
                df = pd.read_excel(file_like_object)
                return DeepRAGExcelParser._dataframe_to_workbook(df)
            except Exception:
                raise ValueError("Spreadsheet input could not be parsed") from None

    @staticmethod
    def _dataframe_to_workbook(df):
        """
        将pandas DataFrame转换为openpyxl工作簿对象

        该方法创建一个新的工作簿，并将DataFrame的数据写入其中。
        第一行为列标题，后续行为数据内容。

        Args:
            df (pandas.DataFrame): 要转换的DataFrame对象

        Returns:
            openpyxl.Workbook: 包含DataFrame数据的工作簿对象
        """
        # 创建新的工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # 写入列标题（第一行）
        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        # 写入数据行（从第二行开始）
        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        return wb

    def html(self, fnm, chunk_rows=256):
        """
        将Excel文件转换为HTML表格格式

        该方法将Excel文件的每个工作表转换为HTML表格，支持大文件的分块处理。
        每个工作表会被分割成多个HTML表格块，以便于处理和显示。

        Args:
            fnm: 文件名或字节数据
            chunk_rows (int): 每个HTML表格块包含的最大行数，默认256行

        Returns:
            list: 包含HTML表格字符串的列表，每个元素是一个表格块
        """
        # 处理输入文件对象
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = DeepRAGExcelParser._load_excel_to_workbook(file_like_object)
        tb_chunks = []

        # 遍历工作簿中的每个工作表
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue

            # 构建表头行HTML
            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{t.value}</th>"
            tb_rows_0 += "</tr>"

            # 按块处理数据行，避免单个表格过大
            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0

                # 处理当前块的数据行
                for r in list(
                    rows[1 + chunk_i * chunk_rows: 1 + (chunk_i + 1) * chunk_rows]
                ):
                    tb += "<tr>"
                    for c in r:  # 移除未使用的变量i
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def __call__(self, fnm):
        """
        将Excel文件转换为结构化文本格式

        该方法是类的主要调用接口，将Excel文件的内容转换为易于阅读的
        结构化文本格式。每行数据会被格式化为"列名：值"的形式，
        多个字段用分号分隔。

        Args:
            fnm: 文件名或字节数据

        Returns:
            list: 包含格式化文本行的列表，每个元素代表Excel中的一行数据
        """
        # 处理输入文件对象
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = DeepRAGExcelParser._load_excel_to_workbook(file_like_object)

        res = []
        # 遍历工作簿中的每个工作表
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue

            # 获取表头行（列名）
            ti = list(rows[0])

            # 处理数据行（从第二行开始）
            for r in list(rows[1:]):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    # 构建"列名：值"格式的字段
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)

                # 用分号连接所有字段
                line = "; ".join(fields)

                # 如果工作表名不是默认的"Sheet"格式，添加工作表名标识
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                res.append(line)
        return res

    @staticmethod
    def row_number(fnm, binary):
        """
        统计文件的总行数

        该方法根据文件扩展名判断文件类型，并统计相应的行数。
        对于Excel文件，统计所有工作表的总行数；
        对于CSV/TXT文件，统计文本行数。

        Args:
            fnm (str): 文件名，用于判断文件类型
            binary (bytes): 文件的二进制内容

        Returns:
            int: 文件的总行数，如果文件类型不支持则返回None
        """
        # 处理Excel文件（.xls, .xlsx等）
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = DeepRAGExcelParser._load_excel_to_workbook(BytesIO(binary))
            total = 0
            # 统计所有工作表的行数
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
            return total

        # 处理CSV和TXT文件
        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            # 检测文件编码
            encoding = find_codec(binary)
            # 解码文件内容并统计行数
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))
